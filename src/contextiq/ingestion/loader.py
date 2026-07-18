"""Document loading with Docling fallback behavior."""

from __future__ import annotations

import hashlib
import logging
from pathlib import Path
from typing import TYPE_CHECKING

from contextiq.ingestion.chunking import DocumentChunker
from contextiq.ingestion.extractors.base import Extractor
from contextiq.ingestion.extractors.docling_standard import DoclingStandardExtractor
from contextiq.ingestion.models import BlockType, DocumentBlock
from contextiq.ingestion.profiles import QUALITY, IngestProfile
from contextiq.ingestion.tree import DocumentTree, TreeBuilder
from contextiq.ingestion.tree_store import TreeStore

if TYPE_CHECKING:
    from contextiq.ingestion.heading_hierarchy import HeadingHierarchyInferencer
    from contextiq.ingestion.summarizer import NodeSummarizer

logger = logging.getLogger(__name__)


class DocumentLoader:
    """Load complex documents into citation-preserving blocks."""

    def __init__(
        self,
        strict_docling: bool = False,
        chunker: DocumentChunker | None = None,
        visuals_dir: Path | None = None,
        enable_picture_enrichment: bool | None = None,
        profile: IngestProfile | None = None,
        extractor: Extractor | None = None,
    ) -> None:
        self.strict_docling = strict_docling
        self.chunker = chunker or DocumentChunker()
        self.visuals_dir = visuals_dir or Path("data/processed/visuals")
        self.profile = profile or QUALITY
        if enable_picture_enrichment is None:
            self.enable_picture_enrichment = self.profile.enable_picture_enrichment
        else:
            self.enable_picture_enrichment = enable_picture_enrichment
        self.extractor = extractor or DoclingStandardExtractor(
            profile=self.profile,
            visuals_dir=self.visuals_dir,
            enable_picture_enrichment=self.enable_picture_enrichment,
        )

    def load(self, path: Path, *, page_range: tuple[int, int] | None = None) -> list[DocumentBlock]:
        """Load a document from disk."""

        if not path.exists():
            raise FileNotFoundError(path)

        if path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            return self.chunker.chunk_blocks(self._load_workbook(path))
        if path.suffix.lower() in {".md", ".markdown", ".txt"}:
            return self.chunker.chunk_blocks(self._load_plain_text(path))

        try:
            return self.chunker.chunk_blocks(
                self.extractor.extract(path, page_range=page_range)
            )
        except Exception as exc:
            if self.strict_docling or not self._can_plain_text_fallback(path):
                raise
            return self.chunker.chunk_blocks(
                self._load_plain_text(path, parser_error=str(exc))
            )

    def load_pdf_range(self, path: Path, *, page_range: tuple[int, int]) -> list[DocumentBlock]:
        """Load one page range from a PDF without chunking policy changes."""

        if path.suffix.lower() != ".pdf":
            raise ValueError("load_pdf_range only supports PDF files")
        return self.chunker.chunk_blocks(
            self.extractor.extract(path, page_range=page_range)
        )

    def build_tree(
        self,
        path: Path,
        *,
        page_range: tuple[int, int] | None = None,
        store: TreeStore | None = None,
        summarizer: NodeSummarizer | None = None,
        hierarchy: HeadingHierarchyInferencer | None = None,
    ) -> DocumentTree:
        """Extract a document and build (and persist) its recursive tree.

        Extractors emit flat heading levels; pass a ``hierarchy`` inferencer to
        recover nesting before the tree is built (issue #10).
        """
        blocks = self.extractor.extract(path, page_range=page_range)
        if hierarchy is not None:
            blocks = hierarchy.assign(blocks)
        tree = TreeBuilder().build(blocks)
        if summarizer is not None:
            block_text = {b.block_id: b.text for b in blocks}
            summarizer.summarize(tree, block_text)
        tree.page_count = max(
            (b.page for b in blocks if b.page is not None), default=None
        )
        if store is not None:
            store.save(tree)
        return tree

    def _load_plain_text(self, path: Path, parser_error: str | None = None) -> list[DocumentBlock]:
        text = path.read_text(encoding="utf-8")
        return self._split_markdown(markdown=text, path=path, parser_error=parser_error)

    def _load_workbook(self, path: Path) -> list[DocumentBlock]:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        document_id = self._document_id(path)
        blocks: list[DocumentBlock] = []
        for sheet in workbook.worksheets:
            rows = [
                ["" if value is None else str(value) for value in row]
                for row in sheet.iter_rows(values_only=True)
            ]
            rows = [row for row in rows if any(cell.strip() for cell in row)]
            if not rows:
                continue
            blocks.append(
                DocumentBlock(
                    document_id=document_id,
                    block_id=f"{document_id}:{len(blocks)}",
                    source_path=str(path),
                    page=None,
                    section_path=[sheet.title],
                    block_type=BlockType.TABLE,
                    text=self._rows_to_markdown_table(rows),
                    metadata={
                        "parser": "openpyxl",
                        "sheet_name": sheet.title,
                        "visual_kind": "spreadsheet",
                        "row_count": len(rows),
                        "column_count": max(len(row) for row in rows),
                    },
                )
            )
        workbook.close()
        return blocks

    def _rows_to_markdown_table(self, rows: list[list[str]]) -> str:
        column_count = max(len(row) for row in rows)
        padded = [row + [""] * (column_count - len(row)) for row in rows]
        header = padded[0]
        body = padded[1:]
        lines = [
            "| " + " | ".join(header) + " |",
            "| " + " | ".join(["---"] * column_count) + " |",
        ]
        lines.extend("| " + " | ".join(row) + " |" for row in body)
        return "\n".join(lines)

    def _split_markdown(
        self,
        markdown: str,
        path: Path,
        parser_error: str | None = None,
    ) -> list[DocumentBlock]:
        document_id = self._document_id(path)
        blocks: list[DocumentBlock] = []
        current_heading: list[str] = []
        metadata = {"parser": "plain_text", "parser_error": parser_error}

        chunks = (chunk for chunk in markdown.split("\n\n") if chunk.strip())
        for index, raw_chunk in enumerate(chunks):
            text = raw_chunk.strip()
            block_type = BlockType.HEADING if text.startswith("#") else BlockType.TEXT
            if block_type == BlockType.HEADING:
                current_heading = [text.lstrip("#").strip()]
            blocks.append(
                DocumentBlock(
                    document_id=document_id,
                    block_id=f"{document_id}:{index}",
                    source_path=str(path),
                    page=None,
                    section_path=current_heading.copy(),
                    block_type=block_type,
                    text=text,
                    metadata=metadata,
                )
            )

        return blocks

    def _can_plain_text_fallback(self, path: Path) -> bool:
        return path.suffix.lower() in {".md", ".markdown", ".txt", ".csv", ".json", ".xml"}

    def _document_id(self, path: Path) -> str:
        digest = hashlib.sha256(str(path.resolve()).encode("utf-8")).hexdigest()[:10]
        return f"{path.stem}-{digest}"
